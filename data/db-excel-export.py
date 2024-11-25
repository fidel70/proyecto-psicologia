import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

class ExportadorDB:
    def __init__(self, ruta_db="db_psicologia_clinic.db"):
        self.ruta_db = ruta_db
        self.fecha_export = datetime.now().strftime("%Y%m%d_%H%M%S")
        
    def conectar_db(self):
        """Establece conexión con la base de datos."""
        try:
            return sqlite3.connect(self.ruta_db)
        except sqlite3.Error as e:
            raise Exception(f"Error conectando a la base de datos: {e}")
    
    def obtener_datos_relacionados(self):
        """Obtiene todos los datos relacionados de las tablas."""
        try:
            conn = self.conectar_db()
            
            # Consulta que une las tres tablas
            query = """
            SELECT 
                pac.codigo as codigo_paciente,
                pac.nombre,
                pac.fecha_nacimiento,
                pac.sexo,
                pac.enfermedad,
                pac.observaciones,
                pac.fecha_registro as fecha_registro_paciente,
                pen.codigo as codigo_pensamiento,
                pen.pensamiento,
                pen.fecha_registro as fecha_registro_pensamiento,
                dim.fecha as fecha_dimension,
                dim.cantidad,
                dim.duracion,
                dim.intensidad
            FROM pacientes pac
            LEFT JOIN pensamientos pen ON pac.id = pen.paciente_id
            LEFT JOIN dimensiones dim ON pen.id = dim.pensamiento_id
            ORDER BY pac.codigo, pen.codigo, dim.fecha
            """
            
            # Obtener datos generales relacionados
            df_general = pd.read_sql_query(query, conn)
            
            # Obtener datos individuales de cada tabla
            df_pacientes = pd.read_sql_query("SELECT * FROM pacientes", conn)
            df_pensamientos = pd.read_sql_query("SELECT * FROM pensamientos", conn)
            df_dimensiones = pd.read_sql_query("SELECT * FROM dimensiones", conn)
            
            conn.close()
            
            return {
                'general': df_general,
                'pacientes': df_pacientes,
                'pensamientos': df_pensamientos,
                'dimensiones': df_dimensiones
            }
            
        except Exception as e:
            raise Exception(f"Error obteniendo datos: {e}")
    
    def exportar_excel(self):
        """Exporta los datos a un archivo Excel con formato."""
        try:
            # Obtener todos los datos
            datos = self.obtener_datos_relacionados()
            
            # Nombre del archivo
            nombre_archivo = f"Export_DB_Psicologia_{self.fecha_export}.xlsx"
            
            # Crear Excel writer con formato
            writer = pd.ExcelWriter(nombre_archivo, engine='openpyxl')
            
            # Exportar cada DataFrame a una hoja diferente
            for nombre, df in datos.items():
                # Exportar datos
                df.to_excel(writer, sheet_name=nombre.capitalize(), index=False)
                
                # Obtener la hoja actual
                worksheet = writer.sheets[nombre.capitalize()]
                
                # Formato de cabeceras
                header_fill = PatternFill(start_color='366092', 
                                        end_color='366092',
                                        fill_type='solid')
                                        
                header_font = Font(bold=True, color='FFFFFF')
                
                # Aplicar formato a cabeceras
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Ajustar ancho de columnas
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
                
                # Alinear todas las celdas al centro
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center')
            
            # Guardar archivo
            writer.close()
            
            print(f"Archivo Excel creado exitosamente: {nombre_archivo}")
            
            # Generar resumen
            self.generar_resumen(datos['general'])
            
        except Exception as e:
            raise Exception(f"Error exportando a Excel: {e}")
    
    def generar_resumen(self, df_general):
        """Genera estadísticas básicas de los datos."""
        try:
            resumen = {
                'Total Pacientes': df_general['codigo_paciente'].nunique(),
                'Total Pensamientos': df_general['codigo_pensamiento'].nunique(),
                'Promedio Intensidad': df_general['intensidad'].mean(),
                'Promedio Duración': df_general['duracion'].mean(),
                'Pacientes por Género': df_general['sexo'].value_counts().to_dict(),
                'Registros por Mes': df_general['fecha_dimension'].dt.to_period('M').value_counts().sort_index().to_dict()
            }
            
            print("\nResumen de Datos:")
            print("-" * 50)
            for key, value in resumen.items():
                print(f"{key}: {value}")
            
        except Exception as e:
            print(f"Error generando resumen: {e}")

def exportar_base_datos():
    try:
        exportador = ExportadorDB()
        exportador.exportar_excel()
    except Exception as e:
        print(f"Error en la exportación: {e}")

if __name__ == "__main__":
    exportar_base_datos()
